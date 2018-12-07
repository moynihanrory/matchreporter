import openpyxl
import pandas as pd
from openpyxl import load_workbook

from constants import EX_SECTORS_SHEET_NAME, EX_HALVES_SHEET_NAME, EX_MATCH_SHEET_NAME, EX_KPI_SHEET_NAME
from fileHelper import getOutputFilename, createAbsoluteFilenameAndDirectory
from report import Report


class Report(Report):

    def insertImagesToSheets(self, filename, plotFilenameList):
        wb = load_workbook(filename)

        for file in plotFilenameList:
            img = openpyxl.drawing.image.Image(file)

            if (file.startswith('match')):
                ws = wb[EX_MATCH_SHEET_NAME]
            elif (file.startswith('half')):
                ws = wb[EX_HALVES_SHEET_NAME]
            else:
                ws = wb[EX_SECTORS_SHEET_NAME]

            ws.add_image(img)

        wb.save(filename)


    def generate(self, analysis, mapping, outputDir):
        absoluteFilename = self.createAbsoluteFilename(analysis, outputDir)

        self.createExcelWorkbook(absoluteFilename, analysis, mapping)

    def createExcelWorkbook(self, absoluteFilename, analysis, mapping):
        writer = pd.ExcelWriter(absoluteFilename)

        analysis.match.to_excel(writer, sheet_name=EX_MATCH_SHEET_NAME)
        analysis.halves.to_excel(writer, sheet_name=EX_HALVES_SHEET_NAME)
        analysis.sectors.to_excel(writer, sheet_name=EX_SECTORS_SHEET_NAME)
        frame = pd.DataFrame.from_dict(list(mapping.items()), orient='columns')#pd.DataFrame.from_records(mapping)
        frame.to_excel(writer, sheet_name=EX_KPI_SHEET_NAME)

        writer.save()

    def createAbsoluteFilename(self, analysis, outputDir):
        filename = getOutputFilename(analysis.teams)

        absoluteFilename = createAbsoluteFilenameAndDirectory(filename, outputDir)

        return absoluteFilename


